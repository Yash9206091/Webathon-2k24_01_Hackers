from flask import Flask, render_template, request, redirect
import openpyxl

# Create a new Flask app
app = Flask(__name__)

# Load the Excel workbook
wb = openpyxl.load_workbook('budget.xlsx')

# Select the active worksheet
ws = wb.active

# Define a function to add a new expense to the Excel file
def add_expense(expense):
    # Append a new row to the table
    ws.append(expense)

    # Save the changes
    wb.save('budget.xlsx')

# Define a function to read expenses from the Excel file
def read_expenses():
    # Get the expenses from the table
    expenses = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        expenses.append(row)
    return expenses

# Define a function to calculate the total budget
def calculate_total_budget():
    # Calculate the total budget by summing the 'Amount' column
    total_budget = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_budget += row[2]
    return total_budget

# Define the home page route
@app.route('/')
def home():
    # Read the expenses from the Excel file
    expenses = read_expenses()

    # Calculate the total budget
    total_budget = calculate_total_budget()

    # Render the home page template with the expenses and total budget
    return render_template('index.html', expenses=expenses, total_budget=total_budget)

# Define a route to add a new expense
@app.route('/expenses', methods=['POST'])
def add_expense_route():
    # Get the expense data
    description = request.form['description']
    category = request.form['category']
    amount = float(request.form['amount'])

    # Create a new expense list
    new_expense = [description, category, amount]

    # Add the new expense to the Excel file
    add_expense(new_expense)

    # Redirect to the home page
    return redirect('/')

# Run the app
if __name__ == '__main__':
    app.run()
